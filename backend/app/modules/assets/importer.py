"""Phase 0 importer/exporter — 云文档/Excel ↔ 资产台账(PRD §13.3)。"""

import io

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.modules.assets.matching import build_dept_index, build_user_index, name_key
from app.modules.assets.migration import HEADER_ALIASES, clean_row
from app.modules.assets.models import Asset
from app.modules.assets.service import generate_asset_code

_MUTABLE = (
    "asset_class",
    "brand_model",
    "spec",
    "serial_number",
    "legacy_code",
    "status",
    "owner_name",
    "department_name",
    "purchase_date",
    "purchase_price",
    "remark",
    "scrap_candidate",
)


def _norm_header(h) -> str:
    # Lark/WPS headers may carry newlines/notes, e.g. "配置\n(PC规则:...)".
    return "" if h is None else str(h).strip().splitlines()[0].strip()


def _map_headers(headers: tuple) -> list[str | None]:
    keys: list[str | None] = []
    for h in headers:
        nh = _norm_header(h)
        key = None
        if nh:
            for cn, internal in HEADER_ALIASES.items():
                if nh == cn or nh.startswith(cn) or cn in nh:
                    key = internal
                    break
        keys.append(key)
    return keys


def parse_workbook(content: bytes) -> list[dict]:
    """Parse the ledger sheet → rows keyed by internal names.

    NOT read_only: real Lark/WPS exports omit the sheet <dimension>, which
    makes read_only mode truncate to a single row. We also pick the sheet
    whose header row best matches the known columns (skips 软件/empty sheets).
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)

    best: tuple[int, list, list[str | None]] | None = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            keys = _map_headers(row)
            mapped = sum(1 for k in keys if k)
            if best is None or mapped > best[0]:
                best = (mapped, list(ws.iter_rows(values_only=True)), keys)
            break  # only inspect the first non-empty row as the header

    if best is None or best[0] < 3:
        return []

    _, all_rows, keys = best
    header_seen = False
    out: list[dict] = []
    for raw in all_rows:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        if not header_seen:
            header_seen = True  # skip the header row itself
            continue
        out.append({keys[i]: raw[i] for i in range(min(len(keys), len(raw))) if keys[i]})
    return out


def import_rows(db: Session, rows: list[dict], operator_id: int | None) -> dict:
    created = updated = needs_review = 0
    users = build_user_index(db)
    depts = build_dept_index(db)

    for row in rows:
        cleaned = clean_row(row)
        data = cleaned.data

        # fuzzy owner/department resolve; unresolved -> needs_review.
        if data["owner_name"]:
            uid = users.resolve(data["owner_name"])
            if uid is not None:
                data["owner_user_id"] = uid
            else:
                cleaned.flag("使用人待匹配飞书账号")
        if data["department_name"]:
            did = depts.get(name_key(data["department_name"]))
            if did is not None:
                data["department_id"] = did
            else:
                cleaned.flag("部门待匹配")

        data["needs_review"] = cleaned.needs_review
        code = data.pop("asset_code", None)

        existing: Asset | None = None
        if data.get("serial_number"):
            existing = db.scalar(
                select(Asset).where(Asset.serial_number == data["serial_number"])
            )
        if existing is None and code:
            existing = db.scalar(select(Asset).where(Asset.asset_code == code))

        if existing is not None:
            for k in _MUTABLE:
                if k in data and data[k] is not None:
                    setattr(existing, k, data[k])
            if data.get("owner_user_id"):
                existing.owner_user_id = data["owner_user_id"]
            if data.get("department_id"):
                existing.department_id = data["department_id"]
            existing.needs_review = data["needs_review"]
            updated += 1
        else:
            asset = Asset(**{k: v for k, v in data.items() if k != "needs_review"})
            asset.needs_review = data["needs_review"]
            asset.asset_code = code or generate_asset_code(db, cleaned.prefix)
            asset.created_by = operator_id
            db.add(asset)
            db.flush()
            created += 1

        if data["needs_review"]:
            needs_review += 1

    db.commit()
    return {"created": created, "updated": updated, "needs_review": needs_review,
            "total": len(rows)}


_EXPORT_COLS = [
    ("资产编号", "asset_code"),
    ("类别", "asset_class"),
    ("品牌型号", "brand_model"),
    ("配置", "spec"),
    ("序列号", "serial_number"),
    ("旧编号", "legacy_code"),
    ("状态", "status"),
    ("使用人", "owner_name"),
    ("部门", "department_name"),
    ("采购日期", "purchase_date"),
    ("原值", "purchase_price"),
    ("报废候选", "scrap_candidate"),
    ("待核", "needs_review"),
    ("备注", "remark"),
]


def export_workbook(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "资产台账"
    ws.append([c[0] for c in _EXPORT_COLS])
    for a in db.scalars(
        select(Asset).where(Asset.deleted_at.is_(None)).order_by(Asset.id)
    ):
        ws.append(
            [
                "" if getattr(a, attr) is None else str(getattr(a, attr))
                for _, attr in _EXPORT_COLS
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rematch_dirty(db: Session) -> dict:
    """Re-resolve owner/department on needs_review assets and recompute the flag.

    Run after a contact sync widens the user/department set. Conservative:
    needs_review stays true while any gap remains (unresolved owner/dept, or
    an unrecognised type → OTH prefix).
    """
    scanned = owner_matched = dept_matched = still_review = 0
    users = build_user_index(db)
    depts = build_dept_index(db)
    rows = db.scalars(select(Asset).where(Asset.needs_review.is_(True))).all()
    for a in rows:
        scanned += 1
        if a.owner_user_id is None and a.owner_name:
            uid = users.resolve(a.owner_name)
            if uid is not None:
                a.owner_user_id = uid
                owner_matched += 1
        if a.department_id is None and a.department_name:
            did = depts.get(name_key(a.department_name))
            if did is not None:
                a.department_id = did
                dept_matched += 1

        gap = (
            (a.owner_name and a.owner_user_id is None)
            or (a.department_name and a.department_id is None)
            or a.asset_code.startswith("OTH-")
        )
        a.needs_review = bool(gap)
        if a.needs_review:
            still_review += 1
    db.commit()
    return {
        "scanned": scanned,
        "owner_matched": owner_matched,
        "dept_matched": dept_matched,
        "still_review": still_review,
    }


def stats(db: Session) -> dict:
    total = db.scalar(select(func.count()).select_from(Asset)) or 0
    review = db.scalar(
        select(func.count()).select_from(Asset).where(Asset.needs_review.is_(True))
    ) or 0
    return {"total": total, "needs_review": review}
