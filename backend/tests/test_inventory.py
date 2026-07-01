import uuid
from concurrent.futures import ThreadPoolExecutor

from app.main import app
from fastapi.testclient import TestClient

client = TestClient(app)


def _admin() -> dict:
    email = f"inv-{uuid.uuid4().hex[:8]}@c.com"
    tok = client.post(
        "/api/auth/dev-login", json={"email": email, "role": "it_admin"}
    ).json()["token"]
    return {"Authorization": f"Bearer {tok}"}


def _emp() -> int:
    return client.post(
        "/api/auth/dev-login", json={"email": f"u-{uuid.uuid4().hex[:6]}@c.com"}
    ).json()["user"]["id"]


def _loc(h) -> int:
    return client.post(
        "/api/inventory/locations", json={"name": f"仓-{uuid.uuid4().hex[:5]}"}, headers=h
    ).json()["id"]


def _cat(h) -> int:
    r = client.post(
        "/api/item-categories",
        json={"name": "测试分类", "code": uuid.uuid4().hex[:8].upper()},
        headers=h,
    )
    assert r.status_code == 201, r.text
    return r.json()["id"]


def _sku(h, loc, safety=5) -> dict:
    r = client.post(
        "/api/skus",
        json={"category_id": _cat(h), "name": "测试鼠标", "safety_stock": safety,
              "default_location_id": loc, "unit": "个"},
        headers=h,
    )
    assert r.status_code == 201, r.text
    return r.json()


def test_receive_issue_return_flow():
    h = _admin()
    loc = _loc(h)
    sku = _sku(h, loc)
    sid = sku["id"]
    emp = _emp()

    assert client.post("/api/inventory/receive",
                       json={"sku_id": sid, "quantity": 20}, headers=h).status_code == 200
    assert client.post("/api/inventory/issue",
                       json={"sku_id": sid, "quantity": 3, "user_id": emp},
                       headers=h).status_code == 200
    assert client.post("/api/inventory/return",
                       json={"sku_id": sid, "quantity": 1}, headers=h).status_code == 200

    listing = client.get(f"/api/skus?q={sku['sku_code']}", headers=h).json()["items"][0]
    assert listing["available"] == 18  # 20 - 3 + 1
    assert listing["level"] == "normal"

    txns = client.get(f"/api/skus/{sku['sku_code']}/transactions", headers=h).json()
    assert len(txns) == 3
    # ledger after_quantity must track the running balance
    assert txns[0]["after_quantity"] == 18


def test_oversell_blocked():
    h = _admin()
    loc = _loc(h)
    sku = _sku(h, loc)
    sid = sku["id"]
    emp = _emp()
    client.post("/api/inventory/receive",
                json={"sku_id": sid, "quantity": 5}, headers=h)
    # 6 > 5 available -> rejected, balance untouched
    r = client.post("/api/inventory/issue",
                     json={"sku_id": sid, "quantity": 6, "user_id": emp}, headers=h)
    assert r.status_code == 409
    avail = client.get(f"/api/skus?q={sku['sku_code']}", headers=h).json()["items"][0]
    assert avail["available"] == 5


def test_concurrent_issue_no_oversell():
    """20 parallel issues of 1 unit against stock of 10 → exactly 10 succeed,
    10 get 409, final available == 0. Proves the FOR UPDATE row lock."""
    h = _admin()
    loc = _loc(h)
    sku = _sku(h, loc, safety=0)
    sid = sku["id"]
    emp = _emp()
    client.post("/api/inventory/receive",
                json={"sku_id": sid, "quantity": 10}, headers=h)

    def issue_one(_):
        return client.post(
            "/api/inventory/issue",
            json={"sku_id": sid, "quantity": 1, "user_id": emp}, headers=h,
        ).status_code

    with ThreadPoolExecutor(max_workers=10) as ex:
        codes = list(ex.map(issue_one, range(20)))

    assert codes.count(200) == 10
    assert codes.count(409) == 10
    final = client.get(f"/api/skus?q={sku['sku_code']}", headers=h).json()["items"][0]
    assert final["available"] == 0


def test_low_stock_scan_task():
    from app.modules.inventory.tasks import scan_low_stock

    h = _admin()
    loc = _loc(h)
    sku = _sku(h, loc, safety=10)
    client.post("/api/inventory/receive",
                json={"sku_id": sku["id"], "quantity": 2}, headers=h)
    res = scan_low_stock()
    assert res["low_stock_count"] >= 1
    assert res["pushed"] is False  # no chat id configured in tests


def test_delete_sku_soft_archives():
    h = _admin()
    loc = _loc(h)

    # SKU holding stock — soft-delete still works (archive, no guard)
    sku = _sku(h, loc)
    client.post("/api/inventory/receive",
                json={"sku_id": sku["id"], "quantity": 6}, headers=h)
    assert client.delete(f"/api/skus/{sku['sku_code']}", headers=h).status_code == 200
    # archived out of the listing …
    assert client.get(f"/api/skus?q={sku['sku_code']}", headers=h).json()["total"] == 0
    # … but the ledger history is preserved
    txns = client.get(f"/api/skus/{sku['sku_code']}/transactions", headers=h).json()
    assert len(txns) == 1
    # deleting again → 404
    assert client.delete(f"/api/skus/{sku['sku_code']}", headers=h).status_code == 404


def test_category_deletable_after_sku_soft_deleted():
    h = _admin()
    loc = _loc(h)
    cid = _cat(h)
    sku = client.post(
        "/api/skus",
        json={"category_id": cid, "name": "鼠标", "default_location_id": loc},
        headers=h,
    ).json()
    # category blocked while it has a live SKU
    assert client.delete(f"/api/item-categories/{cid}", headers=h).status_code == 409
    # soft-deleting the SKU drops it from the category count → category deletable
    assert client.delete(f"/api/skus/{sku['sku_code']}", headers=h).status_code == 200
    assert client.delete(f"/api/item-categories/{cid}", headers=h).status_code == 200


def test_inventory_adjust_up_down_and_clear():
    h = _admin()
    loc = _loc(h)
    sku = _sku(h, loc, safety=0)
    sid = sku["id"]
    code = sku["sku_code"]
    client.post("/api/inventory/receive",
                json={"sku_id": sid, "quantity": 10}, headers=h)

    def avail():
        return client.get(f"/api/skus?q={code}", headers=h).json()["items"][0]["available"]

    # 盘亏 down to 3
    assert client.post("/api/inventory/adjust",
                       json={"sku_id": sid, "target_quantity": 3, "remark": "盘亏"},
                       headers=h).status_code == 200
    assert avail() == 3
    # 盘盈 up to 8
    assert client.post("/api/inventory/adjust",
                       json={"sku_id": sid, "target_quantity": 8}, headers=h).status_code == 200
    assert avail() == 8
    # clear to 0
    assert client.post("/api/inventory/adjust",
                       json={"sku_id": sid, "target_quantity": 0}, headers=h).status_code == 200
    assert avail() == 0

    # each adjust wrote an 'adjustment' ledger row
    txns = client.get(f"/api/skus/{code}/transactions", headers=h).json()
    assert len([t for t in txns if t["transaction_type"] == "adjustment"]) == 3

    # no-op (target == current) → 400
    assert client.post("/api/inventory/adjust",
                       json={"sku_id": sid, "target_quantity": 0}, headers=h).status_code == 400
    # negative target → 400
    assert client.post("/api/inventory/adjust",
                       json={"sku_id": sid, "target_quantity": -1}, headers=h).status_code == 400


def test_export_skus_returns_xlsx():
    from io import BytesIO

    from openpyxl import load_workbook

    h = _admin()
    loc = _loc(h)
    sku = _sku(h, loc)
    r = client.get("/api/skus/export", headers=h)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "SKU 编码" in headers and "当前可用" in headers
    # the SKU we just created shows up somewhere
    codes = [row[headers.index("SKU 编码")].value for row in ws.iter_rows(min_row=2)]
    assert sku["sku_code"] in codes


def test_export_transactions_returns_xlsx():
    from io import BytesIO

    from openpyxl import load_workbook

    h = _admin()
    loc = _loc(h)
    sku = _sku(h, loc)
    client.post(
        "/api/inventory/receive",
        json={"sku_id": sku["id"], "quantity": 5}, headers=h,
    )
    r = client.get("/api/inventory/transactions/export", headers=h)
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "SKU 编码" in headers and "类型" in headers

    # empty date range still produces a valid (header-only) xlsx
    r2 = client.get(
        "/api/inventory/transactions/export",
        params={"date_from": "2099-01-01", "date_to": "2099-12-31"},
        headers=h,
    )
    assert r2.status_code == 200
    ws2 = load_workbook(BytesIO(r2.content)).active
    assert ws2.max_row == 1   # header only


def test_stock_movements_are_audited():
    h = _admin()
    loc = _loc(h)
    sku = _sku(h, loc, safety=0)
    client.post("/api/inventory/receive",
                json={"sku_id": sku["id"], "quantity": 5}, headers=h)
    client.post("/api/inventory/adjust",
                json={"sku_id": sku["id"], "target_quantity": 2}, headers=h)
    logs = client.get("/api/audit-logs", headers=h).json()["items"]
    actions = {r["action"] for r in logs if r["resource_id"] == sku["sku_code"]}
    assert "inventory.receive" in actions
    assert "inventory.adjust" in actions


def test_delete_category_blocked_when_nonempty():
    h = _admin()
    loc = _loc(h)
    cid = _cat(h)
    # empty category → deletable
    assert client.delete(f"/api/item-categories/{cid}", headers=h).status_code == 200

    cid2 = _cat(h)
    client.post(
        "/api/skus",
        json={"category_id": cid2, "name": "鼠标", "default_location_id": loc},
        headers=h,
    )
    blocked = client.delete(f"/api/item-categories/{cid2}", headers=h)
    assert blocked.status_code == 409
    assert "物品" in blocked.json()["detail"]


def test_prize_category_is_delete_protected():
    """The 奖品 (JP) category is system-managed (lottery links prizes to it) and
    must not be deletable from the UI."""
    h = _admin()
    cats = client.get("/api/item-categories", headers=h).json()
    jp = next((c for c in cats if c["code"] == "GIFT"), None)
    assert jp is not None  # created by migration
    r = client.delete(f"/api/item-categories/{jp['id']}", headers=h)
    assert r.status_code == 409
    assert "不可删除" in r.json()["detail"]
    # still present
    assert any(c["code"] == "GIFT" for c in client.get("/api/item-categories", headers=h).json())


def test_prize_category_code_is_edit_locked():
    """The 奖品 (GIFT) category's code can't be changed (name still can)."""
    h = _admin()
    cats = client.get("/api/item-categories", headers=h).json()
    gift = next((c for c in cats if c["code"] == "GIFT"), None)
    assert gift is not None
    # changing the code → 409
    r = client.put(f"/api/item-categories/{gift['id']}", json={"code": "PRZ"}, headers=h)
    assert r.status_code == 409
    assert "不可修改" in r.json()["detail"]
    # changing the name is fine; code stays GIFT
    r2 = client.put(f"/api/item-categories/{gift['id']}", json={"name": "奖品池"}, headers=h)
    assert r2.status_code == 200
    assert r2.json()["code"] == "GIFT"


def test_transactions_list_endpoint():
    """The paginated ledger endpoint returns enriched, filterable rows."""
    h = _admin()
    loc = _loc(h)
    sku = _sku(h, loc)
    sid = sku["id"]
    emp = _emp()
    client.post("/api/inventory/receive", json={"sku_id": sid, "quantity": 20}, headers=h)
    client.post(
        "/api/inventory/issue",
        json={"sku_id": sid, "quantity": 3, "user_id": emp},
        headers=h,
    )

    # filter by this sku so the shared DB's other rows don't interfere
    body = client.get("/api/inventory/transactions", params={"sku_id": sid}, headers=h).json()
    assert body["total"] == 2  # one receive + one issue
    types = {r["type"] for r in body["items"]}
    assert {"purchase_in", "issue_out"} <= types
    row = body["items"][0]
    assert row["sku_code"] == sku["sku_code"]
    assert {"created_at", "type_label", "quantity", "before_quantity",
            "after_quantity", "operator", "recipient"} <= row.keys()
    # the 发放 row shows who it went to; the 入库 row has no recipient
    by_type = {r["type"]: r for r in body["items"]}
    assert by_type["issue_out"]["recipient"]  # non-empty (the employee)
    assert by_type["purchase_in"]["recipient"] == ""

    # pagination: limit=1 → one item but total still 2
    page = client.get(
        "/api/inventory/transactions",
        params={"sku_id": sid, "limit": 1, "offset": 0}, headers=h,
    ).json()
    assert len(page["items"]) == 1 and page["total"] == 2


def test_transactions_list_requires_role():
    emp_tok = client.post(
        "/api/auth/dev-login", json={"email": f"e-{uuid.uuid4().hex[:6]}@c.com"}
    ).json()["token"]
    r = client.get(
        "/api/inventory/transactions",
        headers={"Authorization": f"Bearer {emp_tok}"},
    )
    assert r.status_code == 403
